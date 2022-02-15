import openpyxl

wb = openpyxl.load_workbook("10_oct.xlsx")

print(wb.sheetnames)

# print all sheet nam 
for sheet_name in wb.sheetnames:
    print(sheet_name)
    
    
#select active sheet
sheet = wb.active
cell_A1 = sheet["A1"]
cell_B2 = sheet["B2"]
cell_by_coord = sheet.cell(4, 3) #C4

print(f"value cell A1 => {cell_A1.value} // B2 => {cell_B2.value} // {cell_by_coord} => {cell_by_coord.value}")

for row in range (2, 7):
    cell = sheet.cell(row, 2)
    print(cell.value)
    
# cell empty
print(sheet["B7"].value)

#max row and column
print(f"max row {sheet.max_row}")
print(f"max column {sheet.max_column}")