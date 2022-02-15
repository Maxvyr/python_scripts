import openpyxl

# data_only=True => means nor formula only data
wb_oct = openpyxl.load_workbook("10_oct.xlsx", data_only=True)
wb_nov = openpyxl.load_workbook("11_nov.xlsx", data_only=True)
wb_dec = openpyxl.load_workbook("12_dec.xlsx", data_only=True)

def recover_data_from_excel_file(wb, d):
    sheet = wb.active
    for row in range(2, sheet.max_row):
        value_cell = sheet.cell(row, 1).value
        if not value_cell:
            break
        total_sell = sheet.cell(row, 4).value
        if d.get(value_cell):
            d[value_cell].append(total_sell)
        else:
            d[value_cell] = [total_sell]

datas = {}

recover_data_from_excel_file(wb_oct,datas)
recover_data_from_excel_file(wb_nov,datas)
recover_data_from_excel_file(wb_dec,datas)

print(datas)

#create new excel file with data
wb_result = openpyxl.Workbook()
sheet_res = wb_result.active
sheet_res["A1"] = "Article"
sheet_res["B1"] = "Octobre"
sheet_res["C1"] = "Novembre"
sheet_res["D1"] = "Décembre"

row = 2
for i in datas.items():
    print(i)
    name = i[0]
    values = i[1]
    sheet_res.cell(row, 1).value = name
    # column inc
    for j in range(0, len(values)):
        sheet_res.cell(row, 2+j).value = values[j]
    row += 1

#create chart
chart_ref = openpyxl.chart.Reference(sheet_res, min_col=2,  min_row=2, max_col=sheet_res.max_column, max_row=2)
chart_series = openpyxl.chart.Series(chart_ref, title="Total sells apple")
chart = openpyxl.chart.BarChart3D()
chart.title = "Evolution price apple"
chart.append(chart_series)
sheet_res.add_chart(chart, "F4")

wb_result.save("result.xlsx")